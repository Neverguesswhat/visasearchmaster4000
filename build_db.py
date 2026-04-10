"""
Converts LCA_Disclosure_Data_FY2025_Q1.xlsx -> lca.db (SQLite)
Run once: python3 build_db.py
"""
import sqlite3
import openpyxl
from datetime import datetime, date
import sys
import os

XLSX = os.path.join(os.path.dirname(__file__),
    '../../Companies/Neverguesswhat/LCA_Disclosure_Data_FY2025_Q1.xlsx')
DB = os.path.join(os.path.dirname(__file__), 'lca.db')

COLS = {
    'CASE_NUMBER':            0,
    'CASE_STATUS':            1,
    'RECEIVED_DATE':          2,
    'DECISION_DATE':          3,
    'VISA_CLASS':             5,
    'JOB_TITLE':              6,
    'SOC_CODE':               7,
    'SOC_TITLE':              8,
    'FULL_TIME_POSITION':     9,
    'BEGIN_DATE':            10,
    'END_DATE':              11,
    'TOTAL_WORKER_POSITIONS':12,
    'EMPLOYER_NAME':         19,
    'EMPLOYER_CITY':         23,
    'EMPLOYER_STATE':        24,
    'EMPLOYER_POSTAL_CODE':  25,
    'NAICS_CODE':            31,
    'WORKSITE_CITY':         68,
    'WORKSITE_STATE':        70,
    'WORKSITE_POSTAL_CODE':  71,
    'WAGE_RATE_OF_PAY_FROM': 72,
    'WAGE_UNIT_OF_PAY':      74,
    'PW_WAGE_LEVEL':         77,
    'H1B_DEPENDENT':         84,
    'WILLFUL_VIOLATOR':      85,
}

CREATE_SQL = """
CREATE TABLE IF NOT EXISTS lca (
    id INTEGER PRIMARY KEY,
    case_number TEXT,
    case_status TEXT,
    received_date TEXT,
    decision_date TEXT,
    visa_class TEXT,
    job_title TEXT,
    soc_code TEXT,
    soc_title TEXT,
    full_time TEXT,
    begin_date TEXT,
    end_date TEXT,
    total_workers INTEGER,
    employer_name TEXT,
    employer_city TEXT,
    employer_state TEXT,
    employer_zip TEXT,
    naics_code TEXT,
    worksite_city TEXT,
    worksite_state TEXT,
    worksite_zip TEXT,
    wage_from REAL,
    wage_unit TEXT,
    pw_wage_level TEXT,
    h1b_dependent TEXT,
    willful_violator TEXT
);
CREATE INDEX IF NOT EXISTS idx_status     ON lca(case_status);
CREATE INDEX IF NOT EXISTS idx_visa       ON lca(visa_class);
CREATE INDEX IF NOT EXISTS idx_ws_state   ON lca(worksite_state);
CREATE INDEX IF NOT EXISTS idx_soc        ON lca(soc_title);
CREATE INDEX IF NOT EXISTS idx_employer   ON lca(employer_name);
CREATE INDEX IF NOT EXISTS idx_received   ON lca(received_date);
"""

def fmt_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return str(v)[:10]

def fmt_num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',',''))
    except Exception:
        return None

def fmt_int(v):
    if v is None:
        return None
    try:
        return int(float(str(v).replace(',','')))
    except Exception:
        return None

def main():
    if os.path.exists(DB):
        os.remove(DB)

    print(f'Opening {XLSX}...')
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    total = ws.max_row - 1
    print(f'Rows to import: {total:,}')

    con = sqlite3.connect(DB)
    cur = con.cursor()
    for stmt in CREATE_SQL.strip().split(';'):
        if stmt.strip():
            cur.execute(stmt)

    batch = []
    BATCH_SIZE = 5000
    count = 0

    col_indices = list(COLS.values())

    for row_num, row in enumerate(ws.iter_rows(values_only=True)):
        if row_num == 0:
            continue  # skip header

        vals = (
            row[0],          # case_number
            row[1],          # case_status
            fmt_date(row[2]),# received_date
            fmt_date(row[3]),# decision_date
            row[5],          # visa_class
            row[6],          # job_title
            row[7],          # soc_code
            row[8],          # soc_title
            row[9],          # full_time
            fmt_date(row[10]),# begin_date
            fmt_date(row[11]),# end_date
            fmt_int(row[12]),# total_workers
            row[19],         # employer_name
            row[23],         # employer_city
            row[24],         # employer_state
            row[25],         # employer_zip
            row[31],         # naics_code
            row[68],         # worksite_city
            row[70],         # worksite_state
            row[71],         # worksite_zip
            fmt_num(row[72]),# wage_from
            row[74],         # wage_unit
            row[77],         # pw_wage_level
            row[84],         # h1b_dependent
            row[85],         # willful_violator
        )
        batch.append(vals)
        count += 1

        if len(batch) >= BATCH_SIZE:
            cur.executemany('INSERT INTO lca VALUES (NULL' + ',?' * 25 + ')', batch)
            con.commit()
            batch = []
            pct = count / total * 100
            print(f'\r  {count:>9,} / {total:,}  ({pct:.1f}%)', end='', flush=True)

    if batch:
        cur.executemany('INSERT INTO lca VALUES (NULL' + ',?' * 25 + ')', batch)
        con.commit()

    print(f'\r  {count:,} rows imported.                ')

    # Build lookup tables for dropdowns
    cur.execute("CREATE TABLE IF NOT EXISTS meta_visa   AS SELECT DISTINCT visa_class   AS val FROM lca WHERE visa_class   IS NOT NULL ORDER BY val")
    cur.execute("CREATE TABLE IF NOT EXISTS meta_state  AS SELECT DISTINCT worksite_state AS val FROM lca WHERE worksite_state IS NOT NULL ORDER BY val")
    cur.execute("CREATE TABLE IF NOT EXISTS meta_soc    AS SELECT DISTINCT soc_title    AS val FROM lca WHERE soc_title    IS NOT NULL ORDER BY val")
    cur.execute("CREATE TABLE IF NOT EXISTS meta_status AS SELECT DISTINCT case_status  AS val FROM lca WHERE case_status  IS NOT NULL ORDER BY val")
    cur.execute("SELECT COUNT(*) FROM meta_state"); print('States:', cur.fetchone()[0])
    cur.execute("SELECT COUNT(*) FROM meta_soc");   print('SOC titles:', cur.fetchone()[0])
    con.commit()
    con.close()
    print(f'Done. DB saved to {DB}')

if __name__ == '__main__':
    main()
